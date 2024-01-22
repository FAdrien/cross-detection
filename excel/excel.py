#!/usr/bin/python3
# -*- coding: utf-8 -*-

# -------------------------------- Introduction --------------------------------
'''
'''
# ------------------------------------------------------------------------------


# -------------------------------- Bibliothèques --------------------------------
# Pour l'énumération des types de fichiers Excel : 
from enum import Enum

from utils.fichiers import Fichier

# Pour lire, écrire et enregistrer des fichiers Excel (.xlsx, .xlsm, ...) : 
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter, column_index_from_string

# -------------------------------------------------------------------------------



# -------------------------------- Fonctions --------------------------------
def premiere_colonne_non_cachee(_feuille_excel, _lettre_colonne):
	'''
	'''
	numero = column_index_from_string(_lettre_colonne)
	while _feuille_excel.column_dimensions[get_column_letter(numero)].hidden:
		numero += 1

	return get_column_letter(numero)

def colonnes_consecutives_non_cachees(_feuille_excel, _lettre_colonne, _nombre_de_colonnes):
	'''
	'''
	numero, lettres_colonnes = column_index_from_string(_lettre_colonne), []

	for _ in range(_nombre_de_colonnes):
		lettres_colonnes.append(premiere_colonne_non_cachee(_feuille_excel, get_column_letter(numero)))
		numero = column_index_from_string(lettres_colonnes[-1]) + 1

	return lettres_colonnes

def intersection_lignes_colonnes(_lettres_colonnes, _numeros_lignes):
	'''
	'''
	return [lettre + str(numero) for numero in _numeros_lignes for lettre in _lettres_colonnes]
# ---------------------------------------------------------------------------


# -------------------------------- Classes --------------------------------
class Type(Enum):
	'''
	'''

	SEMAINE = 'SEMAINE',
	WEEKEND = 'WEEKEND',
	VACANCES = 'VACANCES',
	REPAS = 'REPAS'

class Couleur(Enum):
	VERT = PatternFill(fgColor = '9BBB59', fill_type = 'solid'),
	ROUGE = PatternFill(fgColor = 'FF0000', fill_type = 'solid')
	BLEU = PatternFill(fgColor = '6D9EEB', fill_type = 'solid')
	GRIS = PatternFill(fgColor = 'A5A5A5', fill_type = 'solid')

class FeuilleExcel:
	'''
	'''

	def __init__(self, _nom, _fichier_excel):
		'''
		'''
		self.nom = _nom
		self.fichier_excel = _fichier_excel
		# TODO : Traiter le cas où la feuille Excel n'est pas présente dans le workbook Excel.
		self.feuille = self.fichier_excel.workbook[self.nom]

	def __getitem__(self, item):
		return self.feuille[item]

class Excel:
	'''
	'''

	def __init__(self, _chemin, _type, _lecture_seule = False):
		'''
		'''

		self.chemin = _chemin
		self.type = _type
		self.lecture_seule = _lecture_seule
		self.ligne = 5 if self.type == Type.SEMAINE else 7

		self.workbook = load_workbook(filename = self.chemin, read_only = self.lecture_seule, keep_vba = True)
		self.feuilles = {nom : FeuilleExcel(nom, self) for nom in self.workbook.sheetnames}

	def feuille(self, _nom_feuille):
		'''
		'''
		return self.feuilles[_nom_feuille]

	def couleur(self, _feuille, _cellule):
		'''
		'''
		return _feuille[_cellule].fill

	def colorier(self, _feuille, _cellule, _couleur):
		'''
		'''
		_feuille[_cellule].fill = _couleur.value

# -------------------------------------------------------------------------


# -------------------------------- Variables --------------------------------
# Correspondance entre le nom des classes et le nom des feuilles Excel : 
FEUILLES_EXCEL_SEMAINE = {'CPES' : 'CPES', 'MPSI' : 'SUP1 MPSI', 'PCSI' : 'SUP2 PCSI', 'MP' : 'SPE1 MP', 'PSI' : 'SPE2 PSI'}
FEUILLES_EXCEL_WEEKEND = {'CPES' : 'WE CPES', 'MPSI' : 'WE SUP1 MPSI', 'PCSI' : 'WE SUP2 PCSI', 'MP' : 'WE SPE1 MP', 'PSI' : 'WE SPE2 PSI'}

TYPE_FICHIER = {Type.SEMAINE : FEUILLES_EXCEL_SEMAINE, Type.WEEKEND : FEUILLES_EXCEL_WEEKEND}

# ---------------------------------------------------------------------------
